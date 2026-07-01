#!/usr/bin/env python3
"""
Simple HTTP server for the Manufacturing Analytics Dashboard
Serves the dashboard and JSON data
"""

import http.server
import socketserver
import os
import sys
import subprocess
sys.stdout.reconfigure(encoding='utf-8')
from pathlib import Path
import json
from openpyxl import load_workbook
from datetime import datetime, date

import hashlib
import secrets
import time

PORT = 8000
subprocess.run(["python", "process_analytics.py"], check=False)

class MyHTTPRequestHandler(http.server.SimpleHTTPRequestHandler):
    

    def end_headers(self):
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
        self.send_header('Pragma', 'no-cache')
        self.send_header('Expires', '0')
        self.send_header('Access-Control-Allow-Origin', '*')
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_GET(self):
        if self.path.startswith('/api/get-car-entries'):
            self.handle_get_car_entries()
        elif self.path.startswith('/api/get-query-entries'):
            self.handle_get_query_entries()
        elif self.path.startswith('/api/get-maintenance-entries'):
            self.handle_get_maintenance_entries()
        elif self.path.startswith('/api/get-keshkomi-entries'):
            self.handle_get_keshkomi_entries()
        elif self.path.startswith('/api/get-punch-entries'):
            self.handle_get_punch_entries()
        elif self.path.startswith('/api/get-gps-requests'):
            self.handle_get_gps_entries()
        else:
            super().do_GET()

    def do_POST(self):
        if self.path == '/api/login':
            self.handle_login()
        elif self.path == '/api/logout':
            self.handle_logout()
        elif self.path == '/api/forgot-password':
            self.handle_forgot_password()
        elif self.path == '/api/update-safety':
            self.handle_update_safety()
        elif self.path == '/api/update-yearly-target':
            self.handle_update_yearly_target()
        elif self.path == '/api/add-car-entry':
            self.handle_add_car_entry()
        elif self.path == '/api/update-car-actual':
            self.handle_update_car_actual()
        elif self.path == '/api/add-query-entry':
            self.handle_add_query_entry()
        elif self.path == '/api/update-query-stage':
            self.handle_update_query_stage()
        elif self.path == '/api/update-query-summary':
            self.handle_update_query_summary()
        elif self.path == '/api/update-query-milestone':
            self.handle_update_query_milestone()
        elif self.path == '/api/add-maintenance-entry':
            self.handle_add_maintenance_entry()
        elif self.path == '/api/update-maintenance-entry':
            self.handle_update_maintenance_entry()
        elif self.path == '/api/add-keshkomi-entry':
            self.handle_add_keshkomi_entry()
        elif self.path == '/api/update-keshkomi-status':
            self.handle_update_keshkomi_status()
        elif self.path == '/api/add-gps-request':
            self.handle_add_gps_entry()
        elif self.path == '/api/approve-gps-request':
            self.handle_approve_gps_request()
        else:
            self.send_response(404)
            self.end_headers()


    # ── SAFETY ──────────────────────────────────────────────────────
    def handle_update_safety(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            month  = data.get('month')
            status = data.get('status')
            days   = data.get('days')
            lwd_date = data.get('lwd_date', '')
            type_of_month = data.get('type_of_month', '')

            print(f" Safety update received: month={month}, status={status}, days={days}")

            wb = load_workbook('data/Safety_Status.xlsx')
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                cell_val = row[0].value
                if cell_val and str(cell_val).strip().lower() == month.lower():
                    print(f"✅ Match found for {month}, writing values...")
                    if status == 'red':
                        row[1].value = days if days else 1
                        row[2].value = days if days else 1
                    elif status == 'blue':
                        row[1].value = 1
                        row[2].value = 0
                    else:
                        row[1].value = 0
                        row[2].value = 0
                    break

            wb.save('data/Safety_Status.xlsx')
            print(f"Safety Excel saved successfully")

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error in update-safety: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())

    # ── YEARLY TARGET ────────────────────────────────────────────────
    def handle_update_yearly_target(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)
            new_target = float(data.get('yearly_target', 0))

            wb = load_workbook('data/IHTM_Revenue.xlsx')
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            target_col = next((i + 1 for i, h in enumerate(headers) if h and 'target' in str(h).lower()), None)

            if target_col:
                monthly = new_target / 12
                for row in ws.iter_rows(min_row=2):
                    if row[0].value:
                        row[target_col - 1].value = round(monthly, 2)

            wb.save('data/IHTM_Revenue.xlsx')
            print(f"💾 Yearly target updated to {new_target} M")

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error updating yearly target: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())

    # ── CAR USAGE — ADD ─────────────────────────────────────────────
    def handle_add_car_entry(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            wb = load_workbook('data/Car_Usage.xlsx')
            ws = wb.active

            next_row = ws.max_row + 1

            ws.cell(next_row, 1).value = int(data.get('sl_no')) if data.get('sl_no') else next_row - 1
            ws.cell(next_row, 2).value = data.get('date')
            ws.cell(next_row, 3).value = data.get('out_time')
            ws.cell(next_row, 4).value = data.get('tentative')
            ws.cell(next_row, 5).value = data.get('actual')
            ws.cell(next_row, 6).value = data.get('project')
            ws.cell(next_row, 7).value = data.get('name')
            ws.cell(next_row, 8).value = data.get('tm_no')
           
            wb.save('data/Car_Usage.xlsx')
            print(f"Car entry saved: {data.get('name')} on {data.get('date')}")
            wb2 = load_workbook('data/Car_Usage.xlsx')

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error saving car entry: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())

    # ── CAR USAGE — UPDATE ACTUAL RETURN TIME ────────────────────────
    def handle_update_car_actual(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)
            sl_no        = data.get('sl_no')
            actual_return = data.get('actual_return')

            print(f"📝 Actual return update received: Sl No={sl_no}, actual_return={actual_return}")

            wb = load_workbook('data/Car_Usage.xlsx')
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                if str(row[0].value) == str(sl_no):
                    row[4].value = actual_return
                    print(f"✅ Match found for Sl No {sl_no}, updating actual return time...")
                    break

            wb.save('data/Car_Usage.xlsx')
            print(f"💾 Actual return updated for Sl No {sl_no}: {actual_return}")

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error updating actual return: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())

    # ── CAR USAGE — GET ─────────────────────────────────────────────
    def handle_get_car_entries(self):
        try:
            wb = load_workbook('data/Car_Usage.xlsx')
            ws = wb.active
            entries = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    entries.append({
                        'sl_no':     row[0],
                        'date':      row[1].strftime("%d-%m-%Y") if isinstance(row[1], (datetime, date)) else str(row[1]) if row[1] else '',
                        'out_time':  str(row[2]) if row[2] else '',
                        'tentative': str(row[3]) if row[3] else '',
                        'actual':    str(row[4]) if row[4] else '',
                        'project':   row[5],
                        'name':      row[6],
                        'tm_no':     row[7],
                    })

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'entries': entries}).encode())

        except Exception as e:
            print(f"❌ Error reading car entries: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'entries': [], 'error': str(e)}).encode())

    # ── QUERY MONITORING — ADD ──────────────────────────────────────
    def handle_add_query_entry(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            wb = load_workbook('data/Query_Details.xlsx')
            ws = wb.active

            next_row = ws.max_row + 1

            ws.cell(next_row, 1).value = data.get('project_name')
            ws.cell(next_row, 2).value = data.get('customer_name')
            ws.cell(next_row, 3).value = data.get('plant')
            ws.cell(next_row, 4).value = data.get('leader')
            ws.cell(next_row, 5).value = data.get('tool_no')
            ws.cell(next_row, 6).value = data.get('summary')

            # Frontend sends the newly-created project's starting stage as
            # current_stage / current_stage_date (not date_query directly),
            # so stamp that date into the matching milestone column.
            stage_col = {'Query': 7, 'Speculation': 8, 'Concept': 9, 'Estimation': 10, 'PO/Budget': 11}
            start_stage = data.get('current_stage', 'Query')
            start_date  = data.get('current_stage_date', '')
            col = stage_col.get(start_stage, 7)
            ws.cell(next_row, col).value = start_date

            ws.cell(next_row, 12).value = start_date   # current_stage_date
            ws.cell(next_row, 13).value = start_stage  # current_stage

            wb.save('data/Query_Details.xlsx')
            print(f"✅ Query entry saved: {data.get('project_name')}")
            

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error saving query entry: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())

    # ── QUERY MONITORING — GET ──────────────────────────────────────
    def handle_get_query_entries(self):
        try:
            wb = load_workbook('data/Query_Details.xlsx')
            ws = wb.active
            entries = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    entries.append({
                        'project_name':      row[0],
                        'customer_name':     row[1],
                        'plant':             row[2],
                        'leader':            row[3],
                        'tool_no':           row[4],
                        'summary':           row[5] or '',
                        'date_query':        str(row[6])  if row[6]  else '',
                        'date_spec':         str(row[7])  if row[7]  else '',
                        'date_concept':      str(row[8])  if row[8]  else '',
                        'date_estim':        str(row[9])  if row[9]  else '',
                        'date_po':           str(row[10]) if row[10] else '',
                        'current_stage_date': str(row[11]) if row[11] else '',
                        'current_stage':     row[12] if len(row) > 12 and row[12] else '',
                    })

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'entries': entries}).encode())

        except Exception as e:
            print(f"❌ Error reading query entries: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'entries': [], 'error': str(e)}).encode())

    # ── QUERY — UPDATE STAGE + DATE ────────────────────────────────
    def handle_update_query_stage(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            project_name  = data.get('project_name')
            current_stage = data.get('current_stage')
            stage_date    = data.get('stage_date')

            stage_col = {'Query': 7, 'Speculation': 8, 'Concept': 9, 'Estimation': 10, 'PO/Budget': 11}
            col_idx = stage_col.get(current_stage)

            wb = load_workbook('data/Query_Details.xlsx')
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                if row[0].value and str(row[0].value).strip() == str(project_name).strip():
                    if col_idx:
                        row[col_idx - 1].value = stage_date   # date for the selected milestone
                    row[11].value = stage_date                # current_stage_date
                    row[12].value = current_stage              # current_stage
                    break

            wb.save('data/Query_Details.xlsx')
            print(f"✅ Stage updated: {project_name} → {current_stage} on {stage_date}")

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error updating stage: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())

    # ── QUERY — UPDATE SUMMARY ─────────────────────────────────────
    def handle_update_query_summary(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            project_name = data.get('project_name')
            summary      = data.get('summary')

            wb = load_workbook('data/Query_Details.xlsx')
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                if row[0].value and str(row[0].value).strip() == str(project_name).strip():
                    row[5].value = summary
                    break

            wb.save('data/Query_Details.xlsx')
            print(f"✅ Summary updated for {project_name}")

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error updating summary: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())

    # ── QUERY — UPDATE MILESTONE DATE ──────────────────────────────
    def handle_update_query_milestone(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            project_name = data.get('project_name')
            milestone    = data.get('milestone')
            date         = data.get('date')

            milestone_col = {
                'Query':      7,
                'Speculation': 8,
                'Concept':    9,
                'Estimation': 10,
                'PO/Budget':  11,
            }

            col_idx = milestone_col.get(milestone)
            if col_idx is None:
                raise ValueError(f"Unknown milestone: {milestone}")

            wb = load_workbook('data/Query_Details.xlsx')
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                if row[0].value and str(row[0].value).strip() == str(project_name).strip():
                    row[col_idx - 1].value = date
                    break

            wb.save('data/Query_Details.xlsx')
            print(f"✅ Milestone saved: {project_name} — {milestone} = {date}")

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error saving milestone: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())

    # ── CAR MAINTENANCE — ADD ───────────────────────────────────────
    # Columns: 1=week, 2=cleanDate, 3=name, 4=cleanAck, 5=confDate, 6=confAck
    def handle_add_maintenance_entry(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            wb = load_workbook('data/Car_Maintenance_Schedule.xlsx')
            ws = wb.active

            next_row = ws.max_row + 1

            ws.cell(next_row, 1).value = data.get('week')
            ws.cell(next_row, 2).value = data.get('cleanDate')
            ws.cell(next_row, 3).value = data.get('name')
            ws.cell(next_row, 4).value = bool(data.get('cleanAck', False))
            ws.cell(next_row, 5).value = data.get('confDate', '')
            ws.cell(next_row, 6).value = bool(data.get('confAck', False))

            wb.save('data/Car_Maintenance_Schedule.xlsx')
            print(f"✅ Maintenance entry saved: {data.get('name')} on {data.get('cleanDate')}")

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error saving maintenance entry: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())

    # ── CAR MAINTENANCE — UPDATE (generic field/value patch by name) ─
    def handle_update_maintenance_entry(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            name  = data.get('name')
            field = data.get('field')
            value = data.get('value')

            field_col = {
            'week':      1,
            'cleanDate': 2,
            'name':      3,
            'cleanAck':  4,
            'confDate':  5,
            'confAck':   6,
        }
            col_idx = field_col.get(field)
            if col_idx is None:
                raise ValueError(f"Unknown maintenance field: {field}")

            if field in ('cleanAck', 'confAck'):
                value = "Yes" if value else "No"

            wb = load_workbook('data/Car_Maintenance_Schedule.xlsx')
            ws = wb.active

            found = False
            for row in reversed(list(ws.iter_rows(min_row=2))):
                if row[2].value and str(row[2].value).strip().lower() == str(name).strip().lower():
                    row[col_idx - 1].value = value
                    found = True
                    break

            wb.save('data/Car_Maintenance_Schedule.xlsx')
            wb.close()

            print(f"✅ Maintenance entry updated: {name} → {field}={value} (found={found})")

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': found}).encode())

        except Exception as e:
            print(f"❌ Error updating maintenance entry: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())
                           


       # ── CAR MAINTENANCE — GET ───────────────────────────────────────
    def handle_get_maintenance_entries(self):
        try:
            wb = load_workbook('data/Car_Maintenance_Schedule.xlsx')
            ws = wb.active
            entries = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    entries.append({
                        'week':      row[0] or '',
                        'cleanDate': str(row[1]) if row[1] else '',
                        'name':      row[2] or '',
                        'cleanAck':  bool(row[3]) if len(row) > 3 and row[3] is not None else False,
                        'confDate':  str(row[4]) if len(row) > 4 and row[4] else '',
                        'confAck':   bool(row[5]) if len(row) > 5 and row[5] is not None else False,
                    })

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'entries': entries}).encode())

        except Exception as e:
            print(f"❌ Error reading maintenance entries: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'entries': [], 'error': str(e)}).encode())

    # ── KESHKOMI — ADD ──────────────────────────────────────────────
    def handle_add_keshkomi_entry(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            wb = load_workbook('data/KeshKomi.xlsx')
            ws = wb.active

            next_row = ws.max_row + 1

            ws.cell(next_row, 1).value = int(data.get('sl_no')) if data.get('sl_no') else next_row - 1
            ws.cell(next_row, 2).value = data.get('kadai_point')
            ws.cell(next_row, 3).value = data.get('action_plan')
            ws.cell(next_row, 4).value = data.get('responsibility')
            ws.cell(next_row, 5).value = data.get('target_date')
            ws.cell(next_row, 6).value = data.get('status', 'Incomplete')
            ws.cell(next_row, 7).value = bool(data.get('high_priority', False))

            wb.save('data/KeshKomi.xlsx')
            print(f"✅ KeshKomi entry saved: {data.get('kadai_point')}")

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error saving KeshKomi entry: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())

    # ── KESHKOMI — GET ──────────────────────────────────────────────
    def handle_get_keshkomi_entries(self):
        try:
            wb = load_workbook('data/KeshKomi.xlsx')
            ws = wb.active
            entries = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    entries.append({
                        'sl_no':          row[0],
                        'kadai_point':    row[1] or '',
                        'action_plan':    row[2] or '',
                        'responsibility': row[3] or '',
                        'target_date':    str(row[4]) if row[4] else '',
                        'status':         row[5] or 'Incomplete',
                        'high_priority':  bool(row[6]) if len(row) > 6 and row[6] is not None else False,
                    })

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'entries': entries}).encode())

        except Exception as e:
            print(f"❌ Error reading KeshKomi entries: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'entries': [], 'error': str(e)}).encode())

    # ── KESHKOMI — UPDATE STATUS ────────────────────────────────────
    def handle_update_keshkomi_status(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            sl_no = data.get('sl_no')
            new_status = data.get('status')
            high_priority = data.get('high_priority', None)

            wb = load_workbook('data/KeshKomi.xlsx')
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                if row[0].value and str(row[0].value) == str(sl_no):
                    row[5].value = new_status
                    if high_priority is not None:
                        if len(row) > 6:
                            row[6].value = bool(high_priority)
                        else:
                            ws.cell(row[0].row, 7).value = bool(high_priority)
                    break

            wb.save('data/KeshKomi.xlsx')
            print(f"✅ KeshKomi status updated: Sl No {sl_no} → {new_status}")

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error updating KeshKomi status: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())

    # ── PUNCH POINTS — GET ──────────────────────────────────────────
    def handle_get_punch_entries(self):
        try:
            wb = load_workbook('data/Project_Punch_Points.xlsx')
            ws = wb.active
            entries = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    entries.append({
                        'item': row[0] or '',
                        'month': row[1] or '',
                        'punch_points': row[2] or 0,
                    })

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'entries': entries}).encode())

        except Exception as e:
            print(f"❌ Error reading punch entries: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'entries': [], 'error': str(e)}).encode())


    # ── GPS — ADD ──────────────────────────────────────────────
    # Columns: 1=sl_no, 2=request_date, 3=equipment_id, 4=equipment_name,
    #          5=quantity, 6=requested_by, 7=location, 8=remarks,
    #          9=status, 10=approved_by, 11=requester_ack
    def handle_add_gps_entry(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            wb = load_workbook('data/GPS_Request.xlsx')
            ws = wb.active

            next_row = ws.max_row + 1
            sl_no = next_row - 1

            ws.cell(next_row, 1).value = sl_no
            ws.cell(next_row, 2).value = data.get('request_date')
            ws.cell(next_row, 3).value = data.get('part_no')
            ws.cell(next_row, 4).value = data.get('part_name')
            ws.cell(next_row, 5).value = data.get('quantity')
            ws.cell(next_row, 6).value = data.get('requested_by')
            ws.cell(next_row, 7).value = data.get('location')
            ws.cell(next_row, 8).value = data.get('remarks')
            ws.cell(next_row, 9).value = 'Pending'
            ws.cell(next_row, 10).value = ''
            ws.cell(next_row, 11).value = True

            wb.save('data/GPS_Request.xlsx')
            print(f"GPS Request entry saved: {data.get('part_name')}")

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error saving GPS entry: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())

    # ── GPS — GET ──────────────────────────────────────────────
    def handle_get_gps_entries(self):
        try:
            wb = load_workbook('data/GPS_Request.xlsx')
            ws = wb.active
            entries = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    entries.append({
                        'sl_no':          row[0],
                        'request_date':   str(row[1]) if row[1] else '',
                        'part_no':   row[2] or '',
                        'part_name': row[3] or '',
                        'quantity':       row[4] or 0,
                        'requested_by':   row[5] or '',
                        'location':       row[6] or '',
                        'remarks':        row[7] or '',
                        'approved_by':    row[9] if len(row) > 9 and row[9] else '',
                        'requester_ack':  bool(row[10]) if len(row) > 10 and row[10] is not None else False,
                    })

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'entries': entries}).encode())

        except Exception as e:
            print(f"❌ Error reading GPS entries: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'entries': [], 'error': str(e)}).encode())

    # ── GPS — APPROVE ────────────────────────────────────────────
    def handle_approve_gps_request(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body)

            sl_no = data.get('sl_no')
            approved_by = data.get('approved_by', '')

            wb = load_workbook('data/GPS_Request.xlsx')
            ws = wb.active

            for row in ws.iter_rows(min_row=2):
                if row[0].value and str(row[0].value) == str(sl_no):
                    row[8].value = 'Approved'
                    row[9].value = approved_by
                    break

            wb.save('data/GPS_Request.xlsx')
            print(f"✅ GPS request approved: Sl No {sl_no} by {approved_by}")

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': True}).encode())

        except Exception as e:
            print(f"❌ Error approving GPS request: {e}")
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'success': False, 'error': str(e)}).encode())


    def log_message(self, format, *args):
        print(f'[{self.log_date_time_string()}] {format % args}')


def main():
    script_dir = Path(__file__).resolve().parent
    os.chdir(script_dir)

    print(f"""
╔══════════════════════════════════════════════════════════╗
║   IHTM Analytics Dashboard - Development Server         ║
╚══════════════════════════════════════════════════════════╝

📊 Server starting...
🌐 URL: http://localhost:{PORT}
📁 Directory: {script_dir}

Auth (data/users.json — created on first run):
  👤 admin  / admin123  → role: admin  (full edit access)
  👤 viewer / view123   → role: viewer (read-only)

API Endpoints:
  ✓ POST /api/login                    (authenticate)
  ✓ POST /api/logout                   (invalidate session)
  ✓ POST /api/forgot-password          (reset password → prints to console)
  ✓ POST /api/update-safety            (safety calendar → Excel)
  ✓ POST /api/update-yearly-target     (yearly target → Excel)
  ✓ POST /api/add-car-entry            (car usage form → Excel)
  ✓ POST /api/update-car-actual        (update actual return time → Excel)
  ✓ GET  /api/get-car-entries          (read car usage from Excel)
  ✓ POST /api/add-query-entry          (query monitoring form → Excel)
  ✓ GET  /api/get-query-entries        (read query details from Excel)
  ✓ POST /api/update-query-stage       (update query stage → Excel)
  ✓ POST /api/update-query-summary     (update query summary → Excel)
  ✓ POST /api/update-query-milestone   (update query milestone → Excel)
  ✓ POST /api/add-maintenance-entry    (car maintenance form → Excel)
  ✓ POST /api/update-maintenance-entry (update maintenance → Excel)
  ✓ GET  /api/get-maintenance-entries  (read maintenance from Excel)
  ✓ POST /api/add-keshkomi-entry       (KeshKomi form → Excel)
  ✓ GET  /api/get-keshkomi-entries     (read KeshKomi from Excel)
  ✓ POST /api/update-keshkomi-status   (update KeshKomi status → Excel)
  ✓ GET  /api/get-punch-entries        (read punch points from Excel)

⚠️  Press CTRL+C to stop the server

""")

    with socketserver.TCPServer(("", PORT), MyHTTPRequestHandler) as httpd:
        httpd.allow_reuse_address = True # create data/users.json with defaults if missing
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\n🛑 Server stopped.")
            httpd.shutdown()

if __name__ == '__main__':
    main()